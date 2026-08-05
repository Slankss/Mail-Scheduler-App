#!/usr/bin/env python3
"""İstanbul yazılım şirketleri — kurumsal iletişim e-postası toplayıcı.

Ne yapar:
  1. Bir tohum listesindeki (firma adı + web sitesi) her şirketin KENDİ sitesindeki
     iletişim sayfalarını gezer.
  2. Sayfada AÇIKÇA YAYINLANMIŞ kurumsal e-posta adreslerini (info@, iletisim@,
     contact@ ...) çıkarır.
  3. Her şirketten sonra checkpoint yazar; iş yarıda kesilirse kaldığı yerden devam eder.
  4. Sonucu istanbul_yazilim_sirketleri.xlsx dosyasına yazar.

Kurallar (koda gömülü):
  * Adres, şirketin kendi alan adında olmak zorunda. Başka alan adı kabul edilmez.
  * ad.soyad@ biçimindeki kişisel adresler REDDEDİLİR.
  * Hiçbir adres tahmin edilmez/üretilmez. Bulunamazsa hücre boş kalır.

Kullanım:
    python3 arastirma/toplayici.py --tohum arastirma/tohum_sirketler.csv
    python3 arastirma/toplayici.py --devam            # checkpoint'ten devam
    python3 arastirma/toplayici.py --sadece-yaz       # ağa çıkmadan xlsx üret
    python3 arastirma/toplayici.py --dizin-tara "https://clutch.co/tr/profiles/software-development/istanbul"
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from urllib.parse import urljoin, urlparse

import requests

KOK = os.path.dirname(os.path.abspath(__file__))
VARSAYILAN_TOHUM = os.path.join(KOK, "tohum_sirketler.csv")
VARSAYILAN_CHECKPOINT = os.path.join(KOK, "checkpoint.jsonl")
VARSAYILAN_CIKTI = os.path.join(os.path.dirname(KOK), "istanbul_yazilim_sirketleri.xlsx")

BASLIKLAR = ["Firma Adı", "Web Sitesi", "İletişim E-postası", "Sektör / Faaliyet Alanı", "Kaynak"]

UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"

# Denenecek iletişim sayfası yolları — ilk isabette durulur.
YOLLAR = [
    "/", "/iletisim", "/iletisim/", "/tr/iletisim", "/iletisim.html", "/iletisim.php",
    "/contact", "/contact/", "/contact-us", "/contact-us/", "/en/contact",
    "/bize-ulasin", "/hakkimizda", "/about", "/about-us", "/kunye", "/kvkk", "/gizlilik",
]

# Kurumsal (kişiye özel olmayan) yerel adlar — öncelik sırasıyla.
KURUMSAL_YEREL = [
    "info", "iletisim", "contact", "bilgi", "hello", "merhaba", "hi", "mail",
    "kurumsal", "office", "musteri", "musterihizmetleri", "destek", "support", "help",
    "satis", "sales", "bizeulasin", "iletişim", "team", "genel",
]
# Toplanabilir ama daha az tercih edilen kurumsal adresler.
IKINCIL_YEREL = ["kvkk", "ik", "hr", "kariyer", "career", "basin", "press", "pazarlama", "marketing"]

TUM_KURUMSAL = set(KURUMSAL_YEREL) | set(IKINCIL_YEREL)

EPOSTA_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# ad.soyad / ad_soyad / ad-soyad biçimi (kişiye özel) — kurumsal listede yoksa reddedilir.
KISISEL_RE = re.compile(r"^[a-zçğıöşü]+[._\-][a-zçğıöşü]+$", re.IGNORECASE)

COK_PARCALI_TLD = {
    "com.tr", "net.tr", "org.tr", "gov.tr", "edu.tr", "bel.tr", "web.tr", "gen.tr", "av.tr",
    "co.uk", "org.uk", "com.au", "co.jp", "com.br",
}


def kayitli_alan(host: str) -> str:
    """example.com.tr -> example.com.tr ; www.a.example.com -> example.com"""
    host = (host or "").lower().strip()
    if host.startswith("www."):
        host = host[4:]
    parcalar = host.split(".")
    if len(parcalar) >= 3 and ".".join(parcalar[-2:]) in COK_PARCALI_TLD:
        return ".".join(parcalar[-3:])
    return ".".join(parcalar[-2:]) if len(parcalar) >= 2 else host


def site_hostu(site: str) -> str:
    if not site:
        return ""
    if not site.startswith(("http://", "https://")):
        site = "https://" + site
    return urlparse(site).netloc.lower()


def epostalari_ayikla(metin: str, site_alani: str) -> list[str]:
    """Metinden yalnızca şirketin kendi alan adındaki kurumsal adresleri döndürür."""
    bulunan: dict[str, None] = {}
    for ham in EPOSTA_RE.findall(metin or ""):
        adres = ham.strip(" .,:;()<>\"'").lower()
        if adres.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")):
            continue  # dosya adı yakalanmış olabilir
        yerel, _, alan = adres.partition("@")
        if kayitli_alan(alan) != site_alani:
            continue  # başka alan adı — şirketin kendi adresi değil
        if yerel in TUM_KURUMSAL:
            bulunan[adres] = None
            continue
        if KISISEL_RE.match(yerel):
            continue  # ad.soyad@ — kişiye özel, toplanmaz
        # "iletisim2", "info-tr" gibi kurumsal türevler
        cekirdek = re.sub(r"[^a-z]", "", yerel)
        if cekirdek in TUM_KURUMSAL:
            bulunan[adres] = None
    return list(bulunan)


def en_iyi_adres(adresler: list[str]) -> str:
    """info@ > iletisim@ > contact@ ... sırasına göre tek adres seçer."""
    if not adresler:
        return ""
    def puan(a: str) -> int:
        yerel = re.sub(r"[^a-z]", "", a.split("@")[0])
        if yerel in KURUMSAL_YEREL:
            return KURUMSAL_YEREL.index(yerel)
        if yerel in IKINCIL_YEREL:
            return 100 + IKINCIL_YEREL.index(yerel)
        return 500
    return sorted(adresler, key=puan)[0]


def sayfa_getir(oturum: requests.Session, url: str, zaman_asimi: int) -> str:
    try:
        y = oturum.get(url, timeout=zaman_asimi, allow_redirects=True)
        if y.status_code == 200 and "text/html" in y.headers.get("Content-Type", "text/html"):
            return y.text
    except requests.RequestException:
        pass
    return ""


def sirketi_isle(oturum: requests.Session, sirket: dict, zaman_asimi: int, bekleme: float) -> dict:
    """Tek şirketin sitesini gezip kurumsal adres arar."""
    site = (sirket.get("web_sitesi") or "").strip()
    sonuc = dict(sirket)
    sonuc.update({"eposta": "", "eposta_kaynak_url": "", "durum": "site_yok"})
    if not site:
        return sonuc

    host = site_hostu(site)
    alan = kayitli_alan(host)
    taban = f"https://{host}"
    sonuc["durum"] = "site_acilmadi"

    for yol in YOLLAR:
        html = sayfa_getir(oturum, urljoin(taban, yol), zaman_asimi)
        if not html:
            continue
        sonuc["durum"] = "eposta_yok"
        adresler = epostalari_ayikla(html, alan)
        # mailto: bağlantıları
        for m in re.findall(r'mailto:([^"\'?>\s]+)', html, re.IGNORECASE):
            adresler += epostalari_ayikla(m, alan)
        if adresler:
            sonuc["eposta"] = en_iyi_adres(adresler)
            sonuc["eposta_kaynak_url"] = urljoin(taban, yol)
            sonuc["durum"] = "bulundu"
            break
        time.sleep(bekleme)

    return sonuc


# --------------------------------------------------------------------------- IO

def tohum_oku(yol: str) -> list[dict]:
    with open(yol, newline="", encoding="utf-8") as f:
        return [r for r in csv.DictReader(f) if (r.get("firma_adi") or "").strip()]


def checkpoint_oku(yol: str) -> dict[str, dict]:
    kayitlar: dict[str, dict] = {}
    if os.path.exists(yol):
        with open(yol, encoding="utf-8") as f:
            for satir in f:
                satir = satir.strip()
                if not satir:
                    continue
                try:
                    k = json.loads(satir)
                except json.JSONDecodeError:
                    continue
                kayitlar[anahtar(k)] = k
    return kayitlar


def checkpoint_yaz(yol: str, kayit: dict) -> None:
    with open(yol, "a", encoding="utf-8") as f:
        f.write(json.dumps(kayit, ensure_ascii=False) + "\n")


def anahtar(k: dict) -> str:
    """Tekilleştirme anahtarı: önce alan adı, yoksa normalize firma adı."""
    alan = kayitli_alan(site_hostu(k.get("web_sitesi") or ""))
    if alan:
        return alan
    ad = (k.get("firma_adi") or "").lower()
    return re.sub(r"[^a-z0-9]", "", ad)


def xlsx_yaz(satirlar: list[dict], cikti: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Şirketler"

    baslik_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    dolgu = PatternFill("solid", fgColor="1F4E79")
    govde = Font(name="Arial", size=10)

    ws.append(BASLIKLAR)
    for h in ws[1]:
        h.font = baslik_font
        h.fill = dolgu
        h.alignment = Alignment(vertical="center")

    for s in sorted(satirlar, key=lambda r: (r.get("firma_adi") or "").lower()):
        ws.append([
            s.get("firma_adi", ""),
            s.get("web_sitesi", ""),
            s.get("eposta", ""),
            s.get("sektor", ""),
            s.get("kaynak", ""),
        ])

    for satir in ws.iter_rows(min_row=2):
        for h in satir:
            h.font = govde
            h.alignment = Alignment(vertical="center")

    for i, genislik in enumerate([34, 34, 30, 40, 46], start=1):
        ws.column_dimensions[get_column_letter(i)].width = genislik
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    # Açıklama sayfası
    ws2 = wb.create_sheet("OKUBENI")
    for i, satir in enumerate(ACIKLAMA, start=1):
        h = ws2.cell(row=i, column=1, value=satir)
        h.font = Font(name="Arial", size=10, bold=satir.endswith(":") or i == 1)
        h.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.column_dimensions["A"].width = 110

    wb.save(cikti)


ACIKLAMA = [
    "istanbul_yazilim_sirketleri.xlsx — kullanım notu",
    "",
    "İletişim E-postası sütunu:",
    "Yalnızca şirketin KENDİ web sitesinde herkese açık yayınlanmış kurumsal adresler yazılır",
    "(info@, iletisim@, contact@ gibi). Kişiye özel (ad.soyad@) adresler toplanmaz, hiçbir adres",
    "tahmin edilmez. Adres bulunamadıysa hücre BOŞ bırakılır — boş hücre 'yok' değil, 'doğrulanmadı' demektir.",
    "",
    "Doldurma:",
    "Adres sütunu, ağ erişimi olan bir makinede arastirma/toplayici.py çalıştırılarak doldurulur:",
    "    pip install requests openpyxl",
    "    python3 arastirma/toplayici.py --tohum arastirma/tohum_sirketler.csv",
    "Script her şirketten sonra arastirma/checkpoint.jsonl dosyasına yazar; --devam ile kaldığı yerden sürer.",
    "",
    "Kaynak sütunu:",
    "Bilginin hangi siteden alındığını gösterir. 'Ön liste (doğrulanmadı)' yazan satırlar,",
    "istenen kaynak siteleri (TÜBİSAD, startups.watch, Clutch, GoodFirms, LinkedIn) taranamadığı için",
    "başlangıç girdisi olarak eklenmiştir; firma adı ve web sitesi teyide muhtaçtır.",
    "",
    "Gönderim öncesi:",
    "Ticari elektronik ileti göndermeden önce İYS (İleti Yönetim Sistemi) ve KVKK yükümlülüklerini kontrol edin.",
]


# --------------------------------------------------------------------------- dizin tarama

def dizin_tara(url: str, zaman_asimi: int) -> list[dict]:
    """Bir listeleme sayfasından (TÜBİSAD üye listesi, Clutch kategorisi vb.)
    dış firma bağlantılarını + bağlantı metnini çıkarır. Kaba ama kaynaktan bağımsızdır."""
    oturum = requests.Session()
    oturum.headers["User-Agent"] = UA
    html = sayfa_getir(oturum, url, zaman_asimi)
    if not html:
        print(f"[!] Sayfa alınamadı: {url}", file=sys.stderr)
        return []
    kaynak_alani = kayitli_alan(urlparse(url).netloc)
    bulunan: dict[str, dict] = {}
    for eslesme in re.finditer(r'<a[^>]+href="(https?://[^"]+)"[^>]*>(.*?)</a>', html, re.I | re.S):
        hedef, metin = eslesme.group(1), re.sub(r"<[^>]+>", " ", eslesme.group(2))
        metin = re.sub(r"\s+", " ", metin).strip()
        alan = kayitli_alan(urlparse(hedef).netloc)
        if not alan or alan == kaynak_alani:
            continue
        if any(x in alan for x in ("facebook", "twitter", "linkedin", "instagram", "youtube", "google", "x.com")):
            continue
        if alan in bulunan:
            continue
        bulunan[alan] = {
            "firma_adi": metin or alan,
            "web_sitesi": f"https://{alan}",
            "sektor": "",
            "kaynak": url,
        }
    return list(bulunan.values())


# --------------------------------------------------------------------------- ana

def main() -> int:
    p = argparse.ArgumentParser(description="İstanbul yazılım şirketleri iletişim e-postası toplayıcı")
    p.add_argument("--tohum", default=VARSAYILAN_TOHUM, help="Tohum CSV (firma_adi,web_sitesi,sektor,kaynak)")
    p.add_argument("--cikti", default=VARSAYILAN_CIKTI)
    p.add_argument("--checkpoint", default=VARSAYILAN_CHECKPOINT)
    p.add_argument("--devam", action="store_true", help="Checkpoint'te olanları atla")
    p.add_argument("--sadece-yaz", action="store_true", help="Ağa çıkma, mevcut veriden xlsx üret")
    p.add_argument("--dizin-tara", metavar="URL", help="Listeleme sayfasından tohum satırları çıkar ve yazdır")
    p.add_argument("--limit", type=int, default=0, help="En fazla kaç şirket işlensin (0 = hepsi)")
    p.add_argument("--bekleme", type=float, default=1.0, help="İstekler arası bekleme (sn)")
    p.add_argument("--zaman-asimi", type=int, default=15)
    a = p.parse_args()

    if a.dizin_tara:
        satirlar = dizin_tara(a.dizin_tara, a.zaman_asimi)
        yazici = csv.DictWriter(sys.stdout, fieldnames=["firma_adi", "web_sitesi", "sektor", "kaynak"])
        yazici.writeheader()
        yazici.writerows(satirlar)
        print(f"[i] {len(satirlar)} kayıt çıkarıldı.", file=sys.stderr)
        return 0

    tohum = tohum_oku(a.tohum)
    islenmis = checkpoint_oku(a.checkpoint)
    print(f"[i] Tohum: {len(tohum)} şirket · checkpoint: {len(islenmis)} işlenmiş")

    if not a.sadece_yaz:
        oturum = requests.Session()
        oturum.headers["User-Agent"] = UA
        sayac = 0
        for s in tohum:
            k = anahtar(s)
            if a.devam and k in islenmis:
                continue
            if a.limit and sayac >= a.limit:
                break
            sonuc = sirketi_isle(oturum, s, a.zaman_asimi, a.bekleme)
            islenmis[k] = sonuc
            checkpoint_yaz(a.checkpoint, sonuc)   # her şirketten sonra ara kayıt
            sayac += 1
            print(f"  {sonuc['durum']:14s} {s['firma_adi'][:38]:40s} {sonuc['eposta']}")
            time.sleep(a.bekleme)

    # Tekilleştirilmiş nihai satırlar: checkpoint sonucu varsa onu, yoksa tohumu kullan
    nihai: dict[str, dict] = {}
    for s in tohum:
        nihai[anahtar(s)] = dict(s, eposta="")
    for k, v in islenmis.items():
        nihai[k] = v

    xlsx_yaz(list(nihai.values()), a.cikti)
    dolu = sum(1 for v in nihai.values() if v.get("eposta"))
    print(f"[✓] {a.cikti} yazıldı — {len(nihai)} şirket, {dolu} e-posta dolu, {len(nihai) - dolu} boş")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
